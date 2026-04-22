"""
주간보고 취합 Excel 다운로드
GET /api/export/weekly?curr_start=YYYY-MM-DD
관리자 전용
"""
import io
import json
from datetime import datetime, timedelta
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from auth import get_current_user
from database import get_conn

router = APIRouter(prefix="/api/export", tags=["export"])

DAY_KOR = ["월", "화", "수", "목", "금"]


def _month_week_label(d: datetime) -> str:
    fd = datetime(d.year, d.month, 1)
    while fd.weekday() != 0:
        fd += timedelta(days=1)
    wn = max(1, round((d - fd).days / 7) + 1)
    return f"{d.month}월 {wn}주차"


def _thin_border():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="1A1100")
    return Border(left=s, right=s, top=s, bottom=s)


def _med_border():
    from openpyxl.styles import Border, Side
    s = Side(style="medium", color="1A1100")
    return Border(left=s, right=s, top=s, bottom=s)


def _set(cell, value="", bold=False, size=10, color="1A1100",
         bg=None, h="center", v="center", wrap=False, bdr=None):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.value = value
    cell.font = Font(bold=bold, size=size, color=color, name="맑은 고딕")
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if bdr:
        cell.border = bdr


def _apply_border_range(ws, min_row, max_row, min_col, max_col, bdr):
    """범위 내 모든 셀에 테두리 적용 (병합셀 비탑좌 셀 건너뜀)"""
    try:
        from openpyxl.cell import MergedCell
    except ImportError:
        MergedCell = None
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            if MergedCell and isinstance(cell, MergedCell):
                continue
            try:
                cell.border = bdr
            except Exception:
                pass


@router.get("/weekly")
def export_weekly_excel(
    curr_start: str = Query(..., description="이번 주 월요일 (YYYY-MM-DD)"),
    user=Depends(get_current_user),
):
    if user["role"] != "admin":
        raise HTTPException(403, "관리자만 접근 가능합니다")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl 미설치 — 'python -m pip install openpyxl' 후 서버 재시작")

    try:
        return _build_excel(curr_start, user)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(tb)
        raise HTTPException(500, f"Excel 생성 오류: {type(e).__name__}: {e}")


def _build_excel(curr_start: str, user):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 날짜 계산
    mon = datetime.strptime(curr_start, "%Y-%m-%d")
    fri = mon + timedelta(days=4)
    next_mon = mon + timedelta(days=7)
    next_fri = mon + timedelta(days=11)
    curr_end_str   = fri.strftime("%Y-%m-%d")
    prev_start_str = next_mon.strftime("%Y-%m-%d")
    prev_end_str   = next_fri.strftime("%Y-%m-%d")
    week_label = _month_week_label(mon)

    this_week_dates = [mon + timedelta(days=i) for i in range(5)]
    next_week_dates = [next_mon + timedelta(days=i) for i in range(5)]
    all_dates = this_week_dates + next_week_dates

    # DB 조회
    with get_conn() as conn:
        users = conn.execute(
            "SELECT id, name, position FROM users WHERE is_active=1 ORDER BY id"
        ).fetchall()
        reports = conn.execute(
            """SELECT r.*, u.name AS author_name, u.position AS author_position
               FROM reports r JOIN users u ON r.author_id = u.id
               WHERE r.curr_start = ? ORDER BY u.name""",
            (curr_start,),
        ).fetchall()
        schedules = conn.execute(
            "SELECT * FROM schedules WHERE date >= ? AND date <= ?",
            (curr_start, prev_end_str),
        ).fetchall()

    sched = {f"{s['user_id']}_{s['date']}": s["content"] for s in schedules}

    THIN = _thin_border()
    MED  = _med_border()

    # 색상
    C_CREAM    = "FFF8EE"
    C_CREAM_DK = "F5EDDB"
    C_ORANGE   = "FD4401"
    C_INK      = "1A1100"
    C_YELLOW   = "FDCB40"
    C_THIS_BG  = "FFF3E0"   # 금주 헤더 배경
    C_NEXT_BG  = "F3E5F5"   # 차주 헤더 배경

    wb = openpyxl.Workbook()

    # ────────────────────────────────────────────────────────
    # 시트 1: 외부일정표
    # ────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "외부일정표"
    ws1.freeze_panes = "B3"
    ws1.sheet_view.showGridLines = False

    # 행 1 – 이름 / 금주 / 차주 그룹 헤더
    ws1.merge_cells("A1:A2")
    _set(ws1["A1"], "이름", bold=True, size=11, bg=C_CREAM_DK, bdr=THIN)

    ws1.merge_cells("B1:F1")
    _set(ws1["B1"], "금주", bold=True, size=11, color=C_ORANGE, bg=C_THIS_BG, bdr=THIN)

    ws1.merge_cells("G1:K1")
    _set(ws1["G1"], "차주", bold=True, size=11, color="7B1FA2", bg=C_NEXT_BG, bdr=THIN)

    # 행 2 – 날짜 헤더
    for i, d in enumerate(all_dates):
        col = get_column_letter(i + 2)
        bg = C_THIS_BG if i < 5 else C_NEXT_BG
        lbl = f"{d.month}/{d.day}({DAY_KOR[d.weekday()]})"
        _set(ws1[f"{col}2"], lbl, bold=True, size=10, bg=bg, bdr=THIN)

    # 데이터 행
    for ri, u in enumerate(users):
        row = ri + 3
        _set(ws1[f"A{row}"], u["name"], bold=True, size=10, bg=C_CREAM_DK, bdr=THIN)
        for ci, d in enumerate(all_dates):
            col = get_column_letter(ci + 2)
            dt  = d.strftime("%Y-%m-%d")
            content = sched.get(f"{u['id']}_{dt}", "")
            c = ws1[f"{col}{row}"]
            c.value = content
            c.font = Font(size=10, name="맑은 고딕",
                          color=C_INK if content else "CCCCCC")
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = THIN

    # 열 너비 / 행 높이
    ws1.column_dimensions["A"].width = 10
    for i in range(10):
        ws1.column_dimensions[get_column_letter(i + 2)].width = 13
    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[2].height = 22
    for i in range(len(users)):
        ws1.row_dimensions[i + 3].height = 34

    # ────────────────────────────────────────────────────────
    # 시트 2+: 개인 보고서 (한 사람당 1시트)
    # ────────────────────────────────────────────────────────
    def _merge_set(ws, top_left, merge_range, value="", bold=False, size=10,
                   color="1A1100", bg=None, h="center", v="center", wrap=False):
        """병합 후 스타일 설정.
        openpyxl에서 병합셀의 top-left 셀에 4면 테두리를 설정하면
        Excel이 병합 영역 전체의 외곽 테두리로 렌더링함."""
        ws.merge_cells(merge_range)
        c = ws[top_left]
        _set(c, value, bold=bold, size=size, color=color, bg=bg, h=h, v=v,
             wrap=wrap, bdr=THIN)

    def _fmt_items(items):
        """지원 항목 → '1. 제목\n내용\n2. ...' 형식"""
        lines = []
        for i, itm in enumerate(items):
            title   = (itm.get("title") or "").strip()
            content = (itm.get("content") or "").strip()
            if title:
                lines.append(f"{i+1}. {title}")
            if content:
                # 여러 줄이면 들여쓰기
                for ln in content.splitlines():
                    lines.append(f"   {ln}" if title else f"{i+1}. {ln}")
            if title and content and i < len(items) - 1:
                lines.append("")  # 항목 사이 빈줄
        return "\n".join(lines)

    for r in reports:
        name = r["author_name"]
        pos  = r["author_position"] or "사원"
        title = f"{name}_{pos}"[:31]
        ws = wb.create_sheet(title=title)
        ws.sheet_view.showGridLines = False

        curr_sup = json.loads(r["curr_support"]) if r["curr_support"] else []
        prev_sup = json.loads(r["prev_support"]) if r["prev_support"] else []
        todos    = json.loads(r["todo_items"])   if r["todo_items"]   else []
        internal = r["internal_work"] or ""
        shared   = r["shared"]        or ""

        # ── 헤더 영역 (행 1~2) ──────────────────────────────
        _merge_set(ws, "A1", "A1:E2", "주간업무보고", bold=True, size=14, bg=C_CREAM_DK)
        _set(ws["F1"], "전주", bold=True, size=10, bg=C_CREAM_DK, bdr=THIN)
        _merge_set(ws, "G1", "G1:H1", f'{r["curr_start"]} ~ {r["curr_end"]}', size=10)
        _set(ws["I1"], "이름", bold=True, size=10, bg=C_CREAM_DK, bdr=THIN)
        _set(ws["J1"], name, size=10, bdr=THIN)

        _set(ws["F2"], "금주", bold=True, size=10, bg=C_CREAM_DK, bdr=THIN)
        _merge_set(ws, "G2", "G2:H2", f'{r["prev_start"]} ~ {r["prev_end"]}', size=10)
        _set(ws["I2"], "직급", bold=True, size=10, bg=C_CREAM_DK, bdr=THIN)
        _set(ws["J2"], pos, size=10, bdr=THIN)
        # 행 1~2 전체 테두리 보강
        _apply_border_range(ws, 1, 2, 6, 10, THIN)

        # ── 테이블 헤더 (행 3) ───────────────────────────────
        _set(ws["A3"], "구분", bold=True, size=10, bg=C_CREAM_DK, bdr=THIN)
        _merge_set(ws, "B3", "B3:E3",
                   f'전주 업무  ({r["curr_start"]} ~ {r["curr_end"]})',
                   bold=True, size=10, bg=C_CREAM_DK)
        _merge_set(ws, "F3", "F3:J3",
                   f'금주 업무  ({r["prev_start"]} ~ {r["prev_end"]})',
                   bold=True, size=10, bg=C_CREAM_DK)

        # ── 지원 항목 (모든 항목을 하나의 셀에 1. 2. 3. 형식으로) ───────
        SR = 4  # 헤더 바로 아래

        _set(ws[f"A{SR}"], "지원", bold=True, size=10, bg=C_CREAM_DK, bdr=THIN)

        txt1 = _fmt_items(curr_sup) if curr_sup else ""
        _merge_set(ws, f"B{SR}", f"B{SR}:E{SR}", txt1,
                   size=10, h="left", v="top", wrap=True)
        ws[f"B{SR}"].font = Font(size=10, name="맑은 고딕")

        txt2 = _fmt_items(prev_sup) if prev_sup else ""
        _merge_set(ws, f"F{SR}", f"F{SR}:J{SR}", txt2,
                   size=10, h="left", v="top", wrap=True)
        ws[f"F{SR}"].font = Font(size=10, name="맑은 고딕")

        # 행 높이: 항목 수에 비례
        item_lines = max(len(curr_sup), len(prev_sup), 1)
        ws.row_dimensions[SR].height = max(60, item_lines * 28 + 16)

        cur = SR + 1

        # ── todo 항목 ────────────────────────────────────────
        _set(ws[f"A{cur}"], "todo항목\n(일정미정)", bold=True, size=10,
             bg=C_CREAM_DK, wrap=True, bdr=THIN)
        todo_txt = "\n".join([f"{i+1}. {t.get('content','')}" for i, t in enumerate(todos)])
        _merge_set(ws, f"B{cur}", f"B{cur}:J{cur}", todo_txt,
                   size=10, h="left", v="top", wrap=True)
        ws[f"B{cur}"].font = Font(size=10, name="맑은 고딕")
        ws.row_dimensions[cur].height = max(40, len(todos) * 22 + 10)
        cur += 1

        # ── 내부작업 ─────────────────────────────────────────
        _set(ws[f"A{cur}"], "내부작업", bold=True, size=10, bg=C_CREAM_DK, bdr=THIN)
        _merge_set(ws, f"B{cur}", f"B{cur}:J{cur}", internal,
                   size=10, h="left", v="top", wrap=True)
        ws[f"B{cur}"].font = Font(size=10, name="맑은 고딕")
        ws.row_dimensions[cur].height = max(40, internal.count("\n") * 18 + 30)
        cur += 1

        # ── 공유 ─────────────────────────────────────────────
        _set(ws[f"A{cur}"], "공유", bold=True, size=10, bg=C_CREAM_DK, bdr=THIN)
        _merge_set(ws, f"B{cur}", f"B{cur}:J{cur}", shared,
                   size=10, h="left", v="top", wrap=True)
        ws[f"B{cur}"].font = Font(size=10, name="맑은 고딕")
        ws.row_dimensions[cur].height = max(30, shared.count("\n") * 18 + 24)

        # ── 열 너비 ──────────────────────────────────────────
        ws.column_dimensions["A"].width = 12
        for cl in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws.column_dimensions[cl].width = 17
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 26
        ws.row_dimensions[3].height = 22

    # ── 저장 & 반환 ──────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"주간보고_{week_label}_{curr_start}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )
